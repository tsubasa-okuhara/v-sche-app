"""
経費精算・レシート管理 Streamlit アプリケーション
電子帳簿保存法対応

使用方法: streamlit run app.py
"""

import os
import hmac
import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

from database import (
    init_db, insert_receipt, update_receipt, soft_delete_receipt,
    search_receipts, get_receipt_by_id, get_audit_log,
    get_categories, get_unique_vendors, get_receipt_stats
)
from image_utils import validate_image, compute_hash, save_image, load_image_for_display, auto_crop_receipt
from ocr_utils import extract_text, extract_receipt_info

# ===== ページ設定 =====
st.set_page_config(
    page_title="経費精算・レシート管理",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ===== スタイル設定 =====
st.markdown("""
<style>
    .metric-box {
        background-color: #f0f2f6;
        padding: 15px;
        border-radius: 8px;
        margin: 10px 0;
    }
    .success-box {
        background-color: #d4edda;
        padding: 10px;
        border-radius: 5px;
        color: #155724;
    }
    .error-box {
        background-color: #f8d7da;
        padding: 10px;
        border-radius: 5px;
        color: #721c24;
    }
</style>
""", unsafe_allow_html=True)

# ===== 初期化処理 =====
if 'initialized' not in st.session_state:
    init_db()
    st.session_state.initialized = True


# ===== パスワード認証 =====
def _get_app_password() -> str:
    """
    アプリパスワードを取得する優先順位:
    1. 環境変数 APP_PASSWORD
    2. Streamlit secrets の [auth] password
    3. デフォルト (ローカル開発用): 'changeme'
    """
    pw = os.environ.get("APP_PASSWORD")
    if pw:
        return pw
    try:
        if hasattr(st, "secrets") and "auth" in st.secrets:
            secret_pw = st.secrets["auth"].get("password")
            if secret_pw:
                return secret_pw
    except Exception:
        pass
    return "changeme"


def check_password() -> bool:
    """
    パスワード認証。正しく認証されるまで True を返さない。
    認証後は st.session_state['authenticated'] = True が保持される。
    """
    if st.session_state.get("authenticated"):
        return True

    def _verify():
        expected = _get_app_password()
        entered = st.session_state.get("_password_input", "")
        if hmac.compare_digest(entered, expected):
            st.session_state["authenticated"] = True
            st.session_state["_password_input"] = ""
        else:
            st.session_state["_password_wrong"] = True

    st.title("🔐 経費精算・レシート管理")
    st.caption("電子帳簿保存法対応システム")
    st.write("")
    st.info("このアプリを使用するにはパスワードが必要です。")

    st.text_input(
        "パスワード",
        type="password",
        key="_password_input",
        on_change=_verify,
    )

    if st.session_state.get("_password_wrong"):
        st.error("❌ パスワードが違います")

    return False


def format_currency(amount: float) -> str:
    """金額を通貨形式でフォーマット"""
    return f"¥{amount:,.0f}"


def create_excel_download(results: list, date_from=None, date_to=None) -> bytes:
    """検索結果をExcelファイルに変換して返す"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "経費明細"

    # スタイル定義
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    title_font = Font(name="Arial", bold=True, size=14)
    subtitle_font = Font(name="Arial", size=10, color="666666")
    currency_format = '#,##0'
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # タイトル行
    ws.merge_cells("A1:C1")
    ws["A1"] = "経費精算明細書"
    ws["A1"].font = title_font

    # 期間表示
    period_text = ""
    if date_from and date_to:
        period_text = f"期間: {date_from} ～ {date_to}"
    elif date_from:
        period_text = f"期間: {date_from} ～"
    elif date_to:
        period_text = f"期間: ～ {date_to}"
    if period_text:
        ws.merge_cells("A2:C2")
        ws["A2"] = period_text
        ws["A2"].font = subtitle_font

    # 出力日
    ws["A3"] = f"出力日: {datetime.now().strftime('%Y年%m月%d日')}"
    ws["A3"].font = subtitle_font

    # ヘッダー行（5行目から）
    headers = ["No.", "取引日", "金額", "取引先"]
    col_widths = [6, 14, 14, 30]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # データ行
    total_amount = 0
    for row_idx, receipt in enumerate(results, 1):
        row = row_idx + 5
        ws.cell(row=row, column=1, value=row_idx).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=receipt['transaction_date']).border = thin_border
        amount_cell = ws.cell(row=row, column=3, value=receipt['amount'])
        amount_cell.number_format = currency_format
        amount_cell.border = thin_border
        amount_cell.alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=4, value=receipt['vendor']).border = thin_border
        total_amount += receipt['amount']

    # 合計行
    total_row = len(results) + 6
    ws.cell(row=total_row, column=2, value="合計").font = Font(bold=True)
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=2).border = thin_border
    total_cell = ws.cell(row=total_row, column=3, value=total_amount)
    total_cell.font = Font(bold=True)
    total_cell.number_format = currency_format
    total_cell.alignment = Alignment(horizontal="right")
    total_cell.border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="double", color="4472C4"),
        bottom=Side(style="double", color="4472C4"),
    )

    # バッファに書き出し
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ===== ページ1: レシート登録 =====
def page_receipt_registration():
    st.title("📝 レシート登録")

    col1, col2 = st.columns([2, 1])

    with col1:
        st.header("ステップ1: 画像アップロード")
        uploaded_file = st.file_uploader(
            "レシート画像をアップロード（JPG, PNG）",
            type=["jpg", "jpeg", "png"],
            help="ドラッグ&ドロップまたはクリックして選択"
        )

    # ===== 画像バリデーション =====
    image_data = None
    ocr_result = None

    if uploaded_file:
        st.subheader("画像検証結果")

        # ファイルを読み込む
        image_bytes = uploaded_file.getvalue()
        validation = validate_image(uploaded_file)

        col_img, col_check = st.columns([2, 1])

        with col_img:
            # 画像プレビュー
            st.image(image_bytes, caption="アップロード画像", width="stretch")

        with col_check:
            st.write("**画像仕様:**")
            st.write(f"- 形式: {validation['format']}")
            st.write(f"- サイズ: {validation['width']}×{validation['height']}px")
            st.write(f"- 色モード: {validation['color_mode']}")
            st.write(f"- DPI: {validation['dpi']}")

            st.write("**検証:**")

            # DPIチェック
            if validation['dpi'] and validation['dpi'] >= 200:
                st.success(f"✅ DPI要件OK（{validation['dpi']}dpi ≥ 200dpi）")
            else:
                st.error(f"❌ DPI不足（{validation['dpi']}dpi < 200dpi）")

            # 色モードチェック
            if validation['color_mode'] in ['RGB', 'RGBA']:
                st.success(f"✅ 色モードOK（{validation['color_mode']}）")
            else:
                st.error(f"❌ 色モード不正（{validation['color_mode']}）")

            # 全体判定
            if validation['is_valid']:
                st.success("✅ 画像検証: 合格")
            else:
                st.error("❌ 画像検証: 不合格")

        # エラー・警告表示
        if validation['errors']:
            st.error("**エラー:**")
            for error in validation['errors']:
                st.write(f"- {error}")

        if validation['warnings']:
            st.warning("**警告:**")
            for warning in validation['warnings']:
                st.write(f"- {warning}")

        # 検証成功時の処理
        if validation['is_valid']:
            # ===== 自動クロップ =====
            with st.spinner("レシートを自動検出中..."):
                cropped_bytes = auto_crop_receipt(image_bytes)

            if cropped_bytes != image_bytes:
                st.success("✅ レシート部分を自動クロップしました")
                st.image(cropped_bytes, caption="クロップ画像", width="stretch")
                ocr_target_bytes = cropped_bytes
            else:
                ocr_target_bytes = image_bytes

            image_data = {
                'bytes': image_bytes,  # 保存は元画像（法的要件）
                'hash': compute_hash(image_bytes),
                'dpi': validation['dpi'],
                'color_mode': validation['color_mode']
            }

            # ===== OCR抽出 =====
            st.subheader("ステップ2: OCR抽出")

            with st.spinner("テキスト抽出中..."):
                ocr_text = extract_text(ocr_target_bytes, use_mock=False)
                ocr_result = extract_receipt_info(ocr_text)

            st.write("**抽出結果:**")
            st.info(f"""
            - 日付: {ocr_result['date'] or '検出不可'}
            - 金額: {format_currency(ocr_result['amount']) if ocr_result['amount'] else '検出不可'}
            - 店舗: {ocr_result['vendor'] or '検出不可'}
            """)

            if st.checkbox("抽出テキスト全文を表示"):
                st.text_area("OCR抽出テキスト", value=ocr_result['raw_text'], height=200, disabled=True)

    # ===== レシート登録フォーム =====
    if image_data and ocr_result:
        st.header("ステップ3: レシート情報入力")

        with st.form("receipt_form"):
            col1, col2 = st.columns(2)

            with col1:
                transaction_date = st.date_input(
                    "取引年月日",
                    value=datetime.strptime(ocr_result['date'], '%Y-%m-%d').date() if ocr_result['date'] else datetime.now().date(),
                    format="YYYY-MM-DD"
                )

                amount = st.number_input(
                    "金額",
                    value=int(ocr_result['amount']) if ocr_result['amount'] else 0,
                    min_value=0,
                    step=1,
                    format="%d"
                )

            with col2:
                vendors = get_unique_vendors()
                vendor_default = ocr_result['vendor'] if ocr_result['vendor'] and ocr_result['vendor'] in vendors else (vendors[0] if vendors else "")

                vendor = st.selectbox(
                    "取引先",
                    options=vendors + [""] if vendors else [""],
                    index=vendors.index(vendor_default) if vendor_default and vendor_default in vendors else len(vendors),
                    key="vendor_select"
                )

                # 新規取引先入力
                if vendor == "":
                    vendor = st.text_input("取引先を入力", value=ocr_result['vendor'] or "")

                category = st.selectbox(
                    "費目",
                    options=get_categories(),
                    index=0
                )

            description = st.text_area(
                "備考",
                value="",
                height=100
            )

            # 登録ボタン
            submit = st.form_submit_button("レシート登録", use_container_width=True, type="primary")

            if submit:
                # バリデーション
                if not transaction_date:
                    st.error("取引年月日を入力してください")
                elif amount <= 0:
                    st.error("金額を入力してください")
                elif not vendor:
                    st.error("取引先を入力してください")
                elif not category:
                    st.error("費目を選択してください")
                else:
                    try:
                        # レシート登録
                        receipt_id = insert_receipt(
                            transaction_date=transaction_date.isoformat(),
                            amount=float(amount),
                            vendor=vendor,
                            category=category,
                            description=description,
                            image_path=save_image(image_data['bytes'], 0),  # 仮ID
                            image_hash=image_data['hash'],
                            image_dpi=image_data['dpi'],
                            image_color_mode=image_data['color_mode'],
                            ocr_raw_text=ocr_result['raw_text'],
                            created_by="user"
                        )

                        st.success(f"✅ レシート登録完了！（ID: {receipt_id}）")
                        st.balloons()

                        # フォームをリセット
                        st.session_state.clear()

                    except Exception as e:
                        st.error(f"エラー: {str(e)}")


# ===== ページ2: 検索・一覧 =====
def page_search_and_list():
    st.title("🔍 検索・一覧")

    # 検索パネル
    with st.expander("🔎 検索条件", expanded=True):
        col1, col2, col3 = st.columns(3)

        with col1:
            date_from = st.date_input(
                "取引年月日：開始",
                value=datetime.now().date() - timedelta(days=90),
                format="YYYY-MM-DD"
            )

        with col2:
            date_to = st.date_input(
                "取引年月日：終了",
                value=datetime.now().date(),
                format="YYYY-MM-DD"
            )

        with col3:
            vendors = [""] + get_unique_vendors()
            vendor = st.selectbox(
                "取引先",
                options=vendors,
                index=0
            )

        col1, col2 = st.columns(2)

        with col1:
            amount_min = st.number_input(
                "金額：最小",
                value=0,
                min_value=0,
                step=100,
                format="%d"
            )

        with col2:
            amount_max = st.number_input(
                "金額：最大",
                value=999999,
                min_value=0,
                step=100,
                format="%d"
            )

    # 検索実行
    if st.button("検索", use_container_width=True, type="primary"):
        results = search_receipts(
            date_from=date_from.isoformat() if date_from else None,
            date_to=date_to.isoformat() if date_to else None,
            amount_min=amount_min if amount_min > 0 else None,
            amount_max=amount_max if amount_max < 999999 else None,
            vendor=vendor if vendor else None
        )

        if results:
            # 統計情報
            stats = get_receipt_stats(
                date_from=date_from.isoformat(),
                date_to=date_to.isoformat()
            )

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("件数", f"{stats['count']}件")
            with col2:
                st.metric("合計金額", format_currency(stats['total']))
            with col3:
                st.metric("平均金額", format_currency(stats['total'] / stats['count'] if stats['count'] > 0 else 0))

            # 結果テーブル
            st.subheader("検索結果")

            # DataFrameに変換
            df = pd.DataFrame(results)
            df_display = df[[
                'id', 'transaction_date', 'amount', 'vendor', 'category', 'created_at'
            ]].copy()

            df_display.columns = ['ID', '取引日', '金額', '取引先', '費目', '登録日']
            df_display['金額'] = df_display['金額'].apply(lambda x: format_currency(x))

            st.dataframe(
                df_display,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "ID": st.column_config.NumberColumn(width=50),
                    "取引日": st.column_config.TextColumn(width=100),
                    "金額": st.column_config.TextColumn(width=100),
                    "取引先": st.column_config.TextColumn(width=150),
                    "費目": st.column_config.TextColumn(width=100),
                    "登録日": st.column_config.TextColumn(width=150),
                }
            )

            # Excel出力ボタン
            st.subheader("📥 Excel出力")
            excel_data = create_excel_download(
                results,
                date_from=date_from.isoformat() if date_from else None,
                date_to=date_to.isoformat() if date_to else None,
            )
            filename = f"経費明細_{datetime.now().strftime('%Y%m%d')}.xlsx"
            st.download_button(
                label="Excelファイルをダウンロード",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            # 詳細表示用
            st.subheader("詳細確認")
            selected_id = st.selectbox(
                "レシートを選択",
                options=[r['id'] for r in results],
                format_func=lambda x: f"ID: {x}"
            )

            if selected_id:
                st.session_state.selected_receipt_id = selected_id
                st.info("「詳細・編集」ページで詳細を表示・編集できます")

        else:
            st.info("該当するレシートがありません")


# ===== ページ3: 詳細・編集 =====
def page_detail_and_edit():
    st.title("📋 詳細・編集")

    # レシート選択
    col1, col2 = st.columns([3, 1])

    with col1:
        # データベースの全レシートを取得
        all_receipts = search_receipts()
        if all_receipts:
            receipt_id = st.selectbox(
                "レシートを選択",
                options=[r['id'] for r in all_receipts],
                format_func=lambda x: f"ID: {x} - {next((r['transaction_date'] for r in all_receipts if r['id'] == x), 'N/A')}"
            )
        else:
            st.info("登録されたレシートがありません")
            return

    with col2:
        if st.button("🔄 再読み込み"):
            st.rerun()

    # レシート詳細を取得
    receipt = get_receipt_by_id(receipt_id)

    if not receipt:
        st.error("レシートが見つかりません")
        return

    # 画像表示
    st.subheader("レシート画像")

    if receipt['image_path']:
        image = load_image_for_display(receipt['image_path'])
        if image:
            st.image(image, caption="スキャン画像", width="stretch")
        else:
            st.warning("画像が見つかりません")

    # 現在の情報を表示
    st.subheader("登録情報")

    col1, col2 = st.columns(2)

    with col1:
        st.write(f"**ID:** {receipt['id']}")
        st.write(f"**取引日:** {receipt['transaction_date']}")
        st.write(f"**金額:** {format_currency(receipt['amount'])}")

    with col2:
        st.write(f"**取引先:** {receipt['vendor']}")
        st.write(f"**費目:** {receipt['category']}")
        st.write(f"**登録日:** {receipt['created_at']}")

    if receipt['description']:
        st.write(f"**備考:** {receipt['description']}")

    # OCRテキスト表示
    if receipt['ocr_raw_text'] and st.checkbox("OCRテキストを表示"):
        st.text_area("抽出テキスト", value=receipt['ocr_raw_text'], height=150, disabled=True)

    # ===== 編集フォーム =====
    st.divider()
    st.subheader("情報を編集")

    with st.form("edit_form"):
        col1, col2 = st.columns(2)

        with col1:
            new_date = st.date_input(
                "取引年月日",
                value=datetime.strptime(receipt['transaction_date'], '%Y-%m-%d').date(),
                format="YYYY-MM-DD"
            )

            new_amount = st.number_input(
                "金額",
                value=int(receipt['amount']),
                min_value=0,
                step=1,
                format="%d"
            )

        with col2:
            vendors = get_unique_vendors()
            vendor_index = vendors.index(receipt['vendor']) if receipt['vendor'] in vendors else 0

            new_vendor = st.selectbox(
                "取引先",
                options=vendors,
                index=vendor_index
            )

            categories = get_categories()
            category_index = categories.index(receipt['category']) if receipt['category'] in categories else 0

            new_category = st.selectbox(
                "費目",
                options=categories,
                index=category_index
            )

        new_description = st.text_area(
            "備考",
            value=receipt['description'] or "",
            height=100
        )

        change_reason = st.text_input(
            "変更理由 (必須)",
            placeholder="変更内容と理由を記述してください（監査ログに記録されます）",
            help="電子帳簿保存法対応のため、変更理由の記録が必須です"
        )

        col1, col2 = st.columns(2)

        with col1:
            submit_update = st.form_submit_button("変更を保存", use_container_width=True, type="primary")

        with col2:
            submit_delete = st.form_submit_button("論理削除", use_container_width=True, type="secondary")

        if submit_update:
            if not change_reason:
                st.error("変更理由を入力してください")
            else:
                try:
                    update_receipt(
                        receipt_id=receipt_id,
                        transaction_date=new_date.isoformat(),
                        amount=float(new_amount),
                        vendor=new_vendor,
                        category=new_category,
                        description=new_description,
                        reason=change_reason,
                        updated_by="user"
                    )

                    st.success("✅ レシート情報を更新しました")
                    st.rerun()

                except Exception as e:
                    st.error(f"エラー: {str(e)}")

        if submit_delete:
            if not change_reason:
                st.error("削除理由を入力してください")
            else:
                try:
                    soft_delete_receipt(
                        receipt_id=receipt_id,
                        reason=change_reason,
                        deleted_by="user"
                    )

                    st.success("✅ レシートを論理削除しました")
                    st.rerun()

                except Exception as e:
                    st.error(f"エラー: {str(e)}")

    # ===== 監査ログ表示 =====
    st.divider()
    st.subheader("変更履歴")

    audit_logs = get_audit_log(receipt_id)

    if audit_logs:
        audit_df = pd.DataFrame(audit_logs)
        audit_df_display = audit_df[[
            'changed_at', 'action', 'changed_column', 'old_value', 'new_value', 'changed_by', 'reason'
        ]].copy()

        audit_df_display.columns = ['日時', '操作', '項目', '変更前', '変更後', '変更者', '理由']

        st.dataframe(
            audit_df_display,
            use_container_width=True,
            hide_index=True
        )
    else:
        st.info("変更履歴がありません")


# ===== ページ4: 監査ログ =====
def page_audit_log():
    st.title("📊 監査ログ")

    st.write("""
    電子帳簿保存法対応のための変更履歴です。
    全レシートの登録・更新・削除操作が記録されています。
    """)

    # フィルタオプション
    with st.expander("🔎 フィルタ", expanded=False):
        col1, col2 = st.columns(2)

        with col1:
            filter_receipt_id = st.number_input(
                "レシートID（空白で全件表示）",
                value=0,
                min_value=0,
                step=1,
                format="%d"
            )

        with col2:
            filter_action = st.selectbox(
                "操作タイプ",
                options=["", "INSERT", "UPDATE", "DELETE"]
            )

    # ログ取得
    all_logs = get_audit_log()

    # フィルタ適用
    if filter_receipt_id > 0:
        all_logs = [log for log in all_logs if log['receipt_id'] == filter_receipt_id]

    if filter_action:
        all_logs = [log for log in all_logs if log['action'] == filter_action]

    if all_logs:
        st.metric("ログエントリー数", len(all_logs))

        # DataFrameに変換
        log_df = pd.DataFrame(all_logs)
        log_df_display = log_df[[
            'log_id', 'changed_at', 'receipt_id', 'action', 'changed_column',
            'old_value', 'new_value', 'changed_by', 'reason'
        ]].copy()

        log_df_display.columns = ['ログID', '日時', 'レシートID', '操作', '項目', '変更前', '変更後', '変更者', '理由']

        # 色分け（操作タイプ）
        def color_action(action):
            if action == 'INSERT':
                return 'background-color: #d4edda'
            elif action == 'UPDATE':
                return 'background-color: #fff3cd'
            elif action == 'DELETE':
                return 'background-color: #f8d7da'
            return ''

        st.dataframe(
            log_df_display,
            use_container_width=True,
            hide_index=True
        )

    else:
        st.info("ログエントリーがありません")


# ===== メインアプリケーション =====
def main():
    # パスワード認証（失敗時は以降の処理をスキップ）
    if not check_password():
        st.stop()

    # サイドバーナビゲーション
    st.sidebar.title("🧾 経費精算・レシート管理")

    page = st.sidebar.radio(
        "ページを選択",
        options=[
            "📝 レシート登録",
            "🔍 検索・一覧",
            "📋 詳細・編集",
            "📊 監査ログ"
        ],
    )

    st.sidebar.divider()

    # 統計情報をサイドバーに表示
    st.sidebar.subheader("📈 統計")

    stats = get_receipt_stats()
    col1, col2 = st.sidebar.columns(2)

    with col1:
        st.metric("総件数", f"{stats['count']}件")

    with col2:
        st.metric("総金額", format_currency(stats['total']))

    st.sidebar.divider()
    st.sidebar.caption("""
    🔐 電子帳簿保存法対応
    - 全操作が監査ログに記録されます
    - 物理削除は禁止（論理削除のみ）
    - 編集時は理由の記録が必須です
    """)

    st.sidebar.divider()
    if st.sidebar.button("🚪 ログアウト"):
        st.session_state["authenticated"] = False
        st.session_state.pop("_password_wrong", None)
        st.rerun()

    # ページ表示
    if page == "📝 レシート登録":
        page_receipt_registration()
    elif page == "🔍 検索・一覧":
        page_search_and_list()
    elif page == "📋 詳細・編集":
        page_detail_and_edit()
    elif page == "📊 監査ログ":
        page_audit_log()


if __name__ == "__main__":
    main()
